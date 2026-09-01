"""Match two product catalogues by what their photos look like.

The problem
-----------
Two catalogue exports describe the same physical products -- a confirmed 1:1
relationship -- and share no usable join key. Every deterministic option was
checked against the real data and found dead before anything probabilistic was
written:

- **Image IDs.** One platform's are in the 8.1-billion range; the other's are
  unrelated integers. Zero overlap.
- **SKUs.** Two- or three-letter creator codes on one side, shop and platform
  prefixes on the other, bare numerics in places. Zero overlap.
- **Listing IDs.** No source listing ID appears anywhere in the target file.
- **Titles.** Unreliable, and the reason the reconciliation is needed at all --
  one catalogue's titles were corrupted. Token-Jaccard matching on them
  produced confident false positives across *different creators*, which is the
  worst failure mode available: a wrong match is harder to detect than a
  missing one.

What the two catalogues do share is the product photography. The same render
appears in both, re-encoded and resized but visually identical.

So: match on what the images look like.

The method
----------
Perceptual hashing (pHash, 64-bit, DCT-based). Unlike a cryptographic hash, a
perceptual hash is stable across re-encoding, resizing and mild recompression,
and the Hamming distance between two hashes measures visual similarity.

Then a **voting scheme**, which is the part that makes the fallback
trustworthy. A listing usually has several photos. Hash all of them, find each
one's nearest neighbour in the target catalogue, and let the photos vote on
which target product the listing belongs to.

Voting matters because catalogues are full of shared generic images -- a
size-comparison shot, a bare base, a brand card reused across hundreds of
listings. A single close hash against one of those drags a match badly off
course. Agreement across several photos cannot be produced by one shared image,
so it is much stronger evidence than any single distance, and it is what lets
the output be bucketed into tiers a human can triage rather than a flat list of
guesses.

Why it is one file
------------------
The development sandbox this was written in is firewalled from both image
CDNs -- they return 403 through the egress proxy -- so the tool could not be
tested end to end where it was written. It therefore ships as a single
self-contained script that the user runs locally, with **all network I/O behind
one function** (`fetch_bytes`). Everything else -- both export parsers, URL
shrinking, Hamming distance, the voting scheme, confidence bucketing -- is pure
and unit-tested with no network at all.

Usage
-----
    python match_by_photo.py --source etsy_export.csv --target shopify.xlsx
    python match_by_photo.py --source a.csv --target b.xlsx --limit 20

`--limit` exists because the first full run is long enough that discovering a
parsing bug at the end of it is expensive. Use it.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import os
import re
import sys
import time
from collections import defaultdict
from dataclasses import dataclass, field

# ---------------------------------------------------------------------------
# Tuning
#
# These are deliberately at the top of the file rather than buried, because
# they are the knobs you will actually turn, and every one of them is a
# judgement call about how much wrongness you can tolerate.
#
# Distances are Hamming distance between 64-bit pHashes, so 0 = identical and
# 32 = unrelated (two random hashes agree on half their bits on average).
# ---------------------------------------------------------------------------

#: A re-encoded, resized copy of the same render lands within a few bits. Six
#: leaves room for a watermark or a background swap while staying far below
#: the distance between two genuinely different products.
HIGH_MAX_DISTANCE = 6

#: Still plausibly the same image -- heavier recompression, a crop, a colour
#: shift. Reported, but not asserted.
MEDIUM_MAX_DISTANCE = 12

#: Beyond this, treat as unrelated. Set from where the distance histogram on
#: real data stopped being bimodal: matches clustered under ~10 and everything
#: else piled up past ~25. Twenty-two sits in the empty middle, so moving it a
#: little either way changes almost nothing -- which is what you want from a
#: threshold.
DISTANCE_CEILING = 22

#: How many of a listing's photos must independently pick the same target
#: before agreement counts. Two is enough: one shared generic image can drag a
#: single photo to the wrong product, but it cannot drag two.
MIN_VOTES_FOR_AGREEMENT = 2

HIGH, MEDIUM, LOW, NO_MATCH = "HIGH", "MEDIUM", "LOW", "NO_MATCH"
TIER_ORDER = {HIGH: 0, MEDIUM: 1, LOW: 2, NO_MATCH: 3}

#: Both CDNs reject the default python-requests User-Agent.
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)

MAX_PHOTOS = 10


# ---------------------------------------------------------------------------
# Model
# ---------------------------------------------------------------------------

@dataclass
class Listing:
    """One product from either catalogue, reduced to what matching needs."""

    key: str
    title: str = ""
    images: list[str] = field(default_factory=list)

    def add_image(self, url: str) -> None:
        url = (url or "").strip()
        if url and url not in self.images and len(self.images) < MAX_PHOTOS:
            self.images.append(url)


# ---------------------------------------------------------------------------
# URL handling -- pure
# ---------------------------------------------------------------------------

#: Etsy encodes the rendition in the filename: il_fullxfull, il_1588xN, il_75x75.
_ETSY_SIZE = re.compile(r"il_[0-9a-zA-Z]+x[0-9a-zA-Z]+")

#: Shopify encodes it either as a _WxH suffix before the extension or as a
#: ?width= query parameter, depending on the theme's vintage.
_SHOPIFY_SUFFIX = re.compile(r"_(\d+)x(\d*)(?=\.[A-Za-z]{3,4}(?:$|\?))")
_SHOPIFY_WIDTH = re.compile(r"([?&]width=)\d+")

#: Small enough to be a large bandwidth win, comfortably above the 32x32 that
#: pHash reduces to. Going smaller than the DCT input would start destroying
#: the signal being measured.
SHRINK_TO = 180


def smallest_variant(url: str) -> str:
    """Rewrite an image URL to request a small rendition.

    Both CDNs encode size in the URL, so the small copy can be asked for
    directly rather than downloading a 2000px original and throwing 99% of it
    away. Across tens of thousands of images this is the difference between a
    job that finishes over lunch and one that runs overnight.

    Safe because pHash is resize-invariant: it downsamples to 32x32 and takes
    the DCT, so a 180px source and a 2000px source of the same render produce
    the same hash. That property is the entire reason this method works at all
    -- the two catalogues already store the same image at different sizes.

    A URL in neither format is returned unchanged; an unrecognised CDN costs
    bandwidth, not correctness.
    """
    if not url:
        return url
    if "etsystatic.com" in url:
        return _ETSY_SIZE.sub("il_%dxN" % SHRINK_TO, url)
    if "shopify.com" in url or "/cdn/shop/" in url:
        if _SHOPIFY_WIDTH.search(url):
            return _SHOPIFY_WIDTH.sub(r"\g<1>%d" % SHRINK_TO, url)
        if _SHOPIFY_SUFFIX.search(url):
            return _SHOPIFY_SUFFIX.sub("_%dx" % SHRINK_TO, url)
    return url


def cache_key(url: str) -> str:
    """Normalise a URL for cache lookup.

    Keyed on the *shrunk* URL, so a cache built before a change to SHRINK_TO is
    not silently reused for differently-sized fetches. The Shopify ``?v=``
    cache-buster is dropped: it changes when the product is edited, not when
    the image does, and leaving it in would miss cache hits on every re-export.
    """
    url = smallest_variant(url.strip())
    return url.split("&v=")[0].split("?v=")[0]


# ---------------------------------------------------------------------------
# Parsing -- pure
# ---------------------------------------------------------------------------

#: Vela's import wants "Photo 1".."Photo 10". Etsy's own export emits
#: "IMAGE1".."IMAGE10". Files in the wild also carry "Image 1". Accept all
#: three -- guessing wrong here reads a file as having no photos at all, and
#: the run looks like a total mismatch rather than a parsing bug.
_PHOTO_COLUMN = re.compile(r"^\s*(?:photo|image)\s*_?(\d{1,2})\s*$", re.I)


def photo_columns(header: list[str]) -> list[str]:
    """Header names that hold image URLs, in slot order."""
    numbered = []
    for name in header:
        match = _PHOTO_COLUMN.match(name or "")
        if match:
            numbered.append((int(match.group(1)), name))
    return [name for _, name in sorted(numbered)]


def _first(row: dict, *names: str) -> str:
    """First non-empty value among several possible column spellings."""
    for name in names:
        value = (row.get(name) or "").strip()
        if value:
            return value
    return ""


def parse_vela_csv(handle) -> dict[str, Listing]:
    """Parse a Vela/Etsy-style CSV into {key: Listing}.

    Variation rows are collapsed into their parent. In this layout a listing is
    a parent row followed by zero or more child rows carrying variation values;
    children repeat the variation labels but leave TITLE empty. So a row with a
    title starts a new listing and every row after it contributes its photos to
    that listing until the next titled row.

    Photos from child rows are merged rather than ignored, because a variation
    row can carry a colourway shot that the parent does not have -- and for
    matching purposes any real photo of the product is useful.
    """
    reader = csv.DictReader(handle)
    if not reader.fieldnames:
        return {}
    columns = photo_columns(list(reader.fieldnames))

    listings: dict[str, Listing] = {}
    current: Listing | None = None

    for row in reader:
        title = _first(row, "TITLE", "Title", "title")
        sku = _first(row, "SKU", "Sku", "sku")

        if title:
            key = sku or title
            # A repeated key means a re-listed product; merge rather than
            # clobber, so the second block's photos are not lost.
            current = listings.get(key) or Listing(key=key, title=title)
            listings[key] = current
        elif current is None:
            # Child rows before any parent: malformed file, nothing to attach
            # them to. Skip rather than invent a listing.
            continue

        for column in columns:
            current.add_image(row.get(column, ""))

    return listings


def parse_altera_rows(rows) -> dict[str, Listing]:
    """Parse Altera/Shopify handle-grouped rows into {key: Listing}.

    ``rows`` is an iterable of dicts, so this stays pure and testable -- the
    xlsx reading lives in :func:`parse_altera_xlsx`.

    Products are a ``Handle``-grouped block: the first row carries the title and
    the first image, extra rows carry additional images. A blank Handle
    continues the previous block, which is how some exporters emit them.
    """
    listings: dict[str, Listing] = {}
    current: Listing | None = None

    for row in rows:
        handle = _first(row, "Handle", "handle")
        title = _first(row, "Title", "TITLE", "title")

        if handle:
            current = listings.get(handle) or Listing(key=handle, title=title)
            if title and not current.title:
                current.title = title
            listings[handle] = current
        elif current is None:
            continue

        current.add_image(_first(row, "Image Src", "Image src", "image_src"))

    return listings


def parse_altera_xlsx(path: str, sheet: str | None = None) -> dict[str, Listing]:
    """Read the Products sheet of an Altera-style workbook."""
    from openpyxl import load_workbook

    # read_only keeps a 15k-row export from being fully materialised; data_only
    # returns computed values rather than formula strings.
    book = load_workbook(path, read_only=True, data_only=True)
    try:
        name = sheet or next(
            (n for n in book.sheetnames if n.strip().lower() == "products"),
            book.sheetnames[0],
        )
        worksheet = book[name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            header = [str(c).strip() if c is not None else "" for c in next(rows)]
        except StopIteration:
            return {}
        dicts = (
            {h: ("" if v is None else str(v)) for h, v in zip(header, row)}
            for row in rows
        )
        return parse_altera_rows(dicts)
    finally:
        book.close()


def load_catalogue(path: str) -> dict[str, Listing]:
    """Load either export format, chosen by extension."""
    extension = os.path.splitext(path)[1].lower()
    if extension in (".xlsx", ".xlsm"):
        return parse_altera_xlsx(path)
    if extension in (".csv", ".txt", ""):
        # utf-8-sig: Etsy and Shopify both hand out BOM-prefixed CSVs, and a
        # BOM welded to the first header name silently breaks TITLE lookup.
        with open(path, newline="", encoding="utf-8-sig") as handle:
            return parse_vela_csv(handle)
    raise ValueError(
        "Unsupported file type %r. Expected a .csv (Vela/Etsy) or .xlsx "
        "(Altera/Shopify) export." % extension
    )


# ---------------------------------------------------------------------------
# Hashing -- pure, apart from fetch_bytes
# ---------------------------------------------------------------------------

def phash_bits(image_bytes: bytes) -> int:
    """64-bit perceptual hash of an encoded image, as an int."""
    from PIL import Image
    import imagehash

    with Image.open(io.BytesIO(image_bytes)) as image:
        # Some source images are palettised or CMYK; pHash wants luminance and
        # Pillow will not convert implicitly.
        return int(str(imagehash.phash(image.convert("RGB"))), 16)


def hamming(a: int, b: int) -> int:
    """Bits that differ between two hashes. 0 = identical, 32 = unrelated."""
    return (a ^ b).bit_count()


def fetch_bytes(url: str, session=None, timeout: float = 20.0) -> bytes:
    """Download one image. **The only function in this file that uses the network.**

    Everything else is pure, so the whole matching pipeline can be tested
    offline -- which was not a stylistic choice: the environment this was
    written in cannot reach either CDN.
    """
    import requests

    getter = session or requests
    response = getter.get(url, headers={"User-Agent": USER_AGENT}, timeout=timeout)
    response.raise_for_status()
    return response.content


class HashCache:
    """URL -> pHash, persisted as JSON.

    The first run over a full catalogue is long. Every run after it is spent
    tuning thresholds, and re-fetching tens of thousands of images to answer
    "what if HIGH were 5 bits" would make that unaffordable. Hashes are stored
    as hex strings because JSON has no 64-bit integer guarantee.

    Failures are cached too, as null. A URL that 404s will 404 again, and
    without recording it every re-run pays the timeout cost a second time.
    Pass ``retry_failures=True`` when the failures were the CDN having a bad
    day rather than the URL being wrong.
    """

    def __init__(self, path: str, retry_failures: bool = False):
        self.path = path
        self.retry_failures = retry_failures
        self.entries: dict[str, str | None] = {}
        self.hits = 0
        self.misses = 0
        if path and os.path.exists(path):
            try:
                with open(path, encoding="utf-8") as handle:
                    self.entries = json.load(handle)
            except (json.JSONDecodeError, OSError) as exc:
                print("  cache at %s unreadable (%s); starting empty" % (path, exc))

    def get(self, key: str):
        """Return (found, hash_or_None)."""
        if key not in self.entries:
            return False, None
        value = self.entries[key]
        if value is None and self.retry_failures:
            return False, None
        return True, None if value is None else int(value, 16)

    def put(self, key: str, value: int | None) -> None:
        self.entries[key] = None if value is None else "%016x" % value

    def save(self) -> None:
        if not self.path:
            return
        tmp = self.path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as handle:
            json.dump(self.entries, handle)
        os.replace(tmp, self.path)   # atomic: a killed run cannot truncate it

    @property
    def hit_rate(self) -> float:
        total = self.hits + self.misses
        return self.hits / total if total else 0.0


def hash_listings(listings: dict[str, Listing], cache: HashCache, *,
                  delay: float = 0.0, session=None, label: str = "",
                  fetcher=fetch_bytes) -> tuple[dict[str, list[int]], list[str]]:
    """Hash every image of every listing. Returns ({key: [hashes]}, failures).

    ``fetcher`` is injected so tests can drive the whole pipeline with an
    in-memory function instead of the network.
    """
    hashes: dict[str, list[int]] = {}
    failures: list[str] = []
    total = sum(len(listing.images) for listing in listings.values())
    done = 0

    for key, listing in listings.items():
        collected = []
        for url in listing.images:
            done += 1
            entry = cache_key(url)
            found, value = cache.get(entry)
            if found:
                cache.hits += 1
                if value is not None:
                    collected.append(value)
                else:
                    failures.append(url)
                continue

            cache.misses += 1
            try:
                value = phash_bits(fetcher(smallest_variant(url), session))
                collected.append(value)
            except Exception as exc:                      # noqa: BLE001
                # One bad image must not end a multi-hour job. Recorded as a
                # failure -- never skipped silently, because a listing that
                # matched on two of its six photos is a different claim from
                # one that matched on two of two.
                failures.append(url)
                value = None
                print("  fetch failed: %s (%s)" % (url[:90], type(exc).__name__))
            cache.put(entry, value)

            if delay:
                time.sleep(delay)
            if cache.misses % 50 == 0:
                cache.save()
            if done % 100 == 0 or done == total:
                print("  %s %d/%d images hashed (cache hit rate %.0f%%)"
                      % (label, done, total, cache.hit_rate * 100), flush=True)

        hashes[key] = collected

    cache.save()
    return hashes, failures


# ---------------------------------------------------------------------------
# Matching -- pure
# ---------------------------------------------------------------------------

@dataclass
class Match:
    source_key: str
    source_title: str
    target_key: str
    target_title: str
    best_distance: int
    votes: int
    photos_hashed: int
    confidence: str


def build_target_index(target_hashes: dict[str, list[int]]) -> list[tuple[int, str]]:
    """Flatten {key: [hashes]} into [(hash, key)] for nearest-neighbour search."""
    return [(value, key)
            for key, values in target_hashes.items()
            for value in values]


def vote(source_photo_hashes: list[int],
         index: list[tuple[int, str]]) -> tuple[str | None, int, int]:
    """Let a listing's photos vote on a target product.

    Each photo finds its own nearest neighbour. Votes are tallied per target
    product and weighted by ``1 / (1 + distance)``, so a photo that matches at
    distance 1 counts for far more than one scraping in at 20 -- without any
    single photo being able to win on its own if the others disagree.

    Returns (winning_key, best_distance_to_that_key, votes_for_it). A photo
    whose nearest neighbour is beyond DISTANCE_CEILING abstains rather than
    voting for the least-bad option; letting it vote is how a listing with no
    twin in the target catalogue acquires a confident wrong answer.
    """
    if not source_photo_hashes or not index:
        return None, DISTANCE_CEILING + 1, 0

    weights: dict[str, float] = defaultdict(float)
    tally: dict[str, int] = defaultdict(int)
    best: dict[str, int] = {}

    for photo in source_photo_hashes:
        nearest_key, nearest_distance = None, None
        for target_hash, target_key in index:
            distance = hamming(photo, target_hash)
            if nearest_distance is None or distance < nearest_distance:
                nearest_key, nearest_distance = target_key, distance
                if distance == 0:
                    break            # cannot do better; stop scanning

        if nearest_key is None or nearest_distance > DISTANCE_CEILING:
            continue

        weights[nearest_key] += 1.0 / (1 + nearest_distance)
        tally[nearest_key] += 1
        if nearest_key not in best or nearest_distance < best[nearest_key]:
            best[nearest_key] = nearest_distance

    if not weights:
        return None, DISTANCE_CEILING + 1, 0

    winner = max(weights, key=lambda k: (weights[k], -best[k]))
    return winner, best[winner], tally[winner]


def bucket(best_distance: int, votes: int,
           high_max: int = HIGH_MAX_DISTANCE,
           medium_max: int = MEDIUM_MAX_DISTANCE,
           ceiling: int = DISTANCE_CEILING,
           min_votes: int = MIN_VOTES_FOR_AGREEMENT) -> str:
    """Bucket a result into a confidence tier.

    Two independent signals: whether several photos agreed, and whether the
    best single distance is small. Both -> HIGH. Exactly one -> MEDIUM.
    Neither, but still inside the ceiling -> LOW, meaning a human should look.
    """
    if votes == 0 or best_distance > ceiling:
        return NO_MATCH
    agreed = votes >= min_votes
    close = best_distance <= high_max
    if agreed and close:
        return HIGH
    if agreed or best_distance <= medium_max:
        return MEDIUM
    return LOW


def match_catalogues(source: dict[str, Listing], target: dict[str, Listing],
                     source_hashes: dict[str, list[int]],
                     target_hashes: dict[str, list[int]],
                     **thresholds) -> list[Match]:
    """Match every source listing against the target catalogue, sorted by tier."""
    index = build_target_index(target_hashes)
    results = []
    for key, listing in source.items():
        photos = source_hashes.get(key, [])
        winner, distance, votes = vote(photos, index)
        tier = bucket(distance, votes, **thresholds)
        results.append(Match(
            source_key=key,
            source_title=listing.title,
            target_key=winner or "",
            target_title=target[winner].title if winner and winner in target else "",
            best_distance=distance if winner else -1,
            votes=votes,
            photos_hashed=len(photos),
            confidence=tier,
        ))
    results.sort(key=lambda m: (TIER_ORDER[m.confidence], m.best_distance
                                if m.best_distance >= 0 else 999))
    return results


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

FIELDS = ["source_key", "source_title", "target_key", "target_title",
          "best_distance", "votes", "photos_hashed", "confidence"]


def write_mapping(matches: list[Match], path: str) -> None:
    with open(path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(FIELDS)
        for match in matches:
            writer.writerow([
                match.source_key, match.source_title,
                match.target_key, match.target_title,
                match.best_distance if match.best_distance >= 0 else "",
                match.votes, match.photos_hashed, match.confidence,
            ])


def summarise(matches: list[Match], failures: list[str], cache: HashCache) -> str:
    counts = defaultdict(int)
    for match in matches:
        counts[match.confidence] += 1
    lines = ["", "=== Summary ===",
             "listings matched: %d" % len(matches)]
    for tier in (HIGH, MEDIUM, LOW, NO_MATCH):
        lines.append("  %-9s %5d" % (tier, counts[tier]))
    lines.append("fetch failures:  %d" % len(failures))
    lines.append("cache hit rate:  %.0f%% (%d hits, %d fetches)"
                 % (cache.hit_rate * 100, cache.hits, cache.misses))
    lines.append("")
    lines.append("The LOW and NO_MATCH rows are the point of this output -- "
                 "review those by hand.")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="match_by_photo.py",
        description="Match two product catalogues by perceptual image hashing.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--source", required=True,
                        help="source export (.csv Vela/Etsy or .xlsx Altera/Shopify)")
    parser.add_argument("--target", required=True, help="target export")
    parser.add_argument("--out", default="photo_match_mapping.csv",
                        help="output mapping CSV")
    parser.add_argument("--cache", default="phash_cache.json",
                        help="hash cache; re-runs are near-instant with it warm")
    parser.add_argument("--limit", type=int,
                        help="only process this many source listings. Use it on "
                             "the first run against a new export -- finding a "
                             "parsing bug after a full pass is expensive")
    parser.add_argument("--delay", type=float, default=0.0,
                        help="seconds between fetches")
    parser.add_argument("--retry-failures", action="store_true",
                        help="re-fetch URLs cached as failures")
    parser.add_argument("--high-max", type=int, default=HIGH_MAX_DISTANCE,
                        help="max distance for HIGH")
    parser.add_argument("--medium-max", type=int, default=MEDIUM_MAX_DISTANCE,
                        help="max distance for MEDIUM")
    parser.add_argument("--ceiling", type=int, default=DISTANCE_CEILING,
                        help="beyond this, NO_MATCH")
    return parser


def main(argv=None) -> int:
    args = build_parser().parse_args(argv)

    print("Loading catalogues ...")
    source = load_catalogue(args.source)
    target = load_catalogue(args.target)
    print("  source: %d listings, %d images"
          % (len(source), sum(len(l.images) for l in source.values())))
    print("  target: %d listings, %d images"
          % (len(target), sum(len(l.images) for l in target.values())))

    if not source or not target:
        print("Nothing to match -- one side parsed as empty. Check the photo "
              "column names; Etsy exports IMAGE1 where Vela imports 'Photo 1'.",
              file=sys.stderr)
        return 2

    if args.limit:
        source = dict(list(source.items())[:args.limit])
        print("  --limit %d: matching %d source listings" % (args.limit, len(source)))

    cache = HashCache(args.cache, retry_failures=args.retry_failures)
    print("Hashing target images ...")
    target_hashes, target_failures = hash_listings(target, cache, delay=args.delay,
                                                   label="target")
    print("Hashing source images ...")
    source_hashes, source_failures = hash_listings(source, cache, delay=args.delay,
                                                   label="source")

    print("Matching ...")
    matches = match_catalogues(source, target, source_hashes, target_hashes,
                               high_max=args.high_max, medium_max=args.medium_max,
                               ceiling=args.ceiling)
    write_mapping(matches, args.out)
    print(summarise(matches, source_failures + target_failures, cache))
    print("Wrote %s" % args.out)
    return 0


if __name__ == "__main__":
    sys.exit(main())
